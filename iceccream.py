from openpyxl import load_workbook
workbook = load_workbook(filename="C:\Users\Dell\Desktop\icecream.xlsx")
worksheet = workbook.active
lst=[]
for row in worksheet.iter_rows(values_only=True):
    lst.append(row[0])
print(lst)
lst1=[]
for row in worksheet.iter_rows(values_only=True):
    lst1.append(row[1])
print(lst1)
lst2=[]
for row in worksheet.iter_rows(values_only=True):
    lst2.append(row[2])
print(lst2)
input_flavor_3 = input("Enter the flavor 3: ")
matching_index = -1
for i in range(len(lst2)):
    if lst2[i] == input_flavor_3:
        matching_index = i 

if matching_index != -1:
    print("The flavor 1 for", input_flavor_3, "is", lst[matching_index])
else:
    print("No match found for flavor 3:", input_flavor_3)
